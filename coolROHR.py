# ==============================================================================
# coolROHR.py  –  Kältemittel-Rohrdimensionierung für CU-Leitungen
# °coolsulting e.U. | Michael Schäpers | v1.0
# ==============================================================================

import streamlit as st
import math
import numpy as np
import pandas as pd
import io
import base64
import os
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# CU-ROHR TABELLE  (OD × Wandstärke → ID, alle Maße in mm)
# ─────────────────────────────────────────────────────────────────────────────
CU_PIPES = [
    {"od": 10.0,  "wall": 1.0, "id": 8.0,   "label": "10 × 1,0"},
    {"od": 12.0,  "wall": 1.0, "id": 10.0,  "label": "12 × 1,0"},
    {"od": 15.0,  "wall": 1.0, "id": 13.0,  "label": "15 × 1,0"},
    {"od": 18.0,  "wall": 1.0, "id": 16.0,  "label": "18 × 1,0"},
    {"od": 22.0,  "wall": 1.0, "id": 20.0,  "label": "22 × 1,0"},
    {"od": 28.0,  "wall": 1.5, "id": 25.0,  "label": "28 × 1,5"},
    {"od": 35.0,  "wall": 1.5, "id": 32.0,  "label": "35 × 1,5"},
    {"od": 42.0,  "wall": 1.5, "id": 39.0,  "label": "42 × 1,5"},
    {"od": 54.0,  "wall": 2.0, "id": 50.0,  "label": "54 × 2,0"},
    {"od": 64.0,  "wall": 2.0, "id": 60.0,  "label": "64 × 2,0"},
    {"od": 76.1,  "wall": 2.0, "id": 72.1,  "label": "76,1 × 2,0"},
    {"od": 88.9,  "wall": 2.5, "id": 83.9,  "label": "88,9 × 2,5"},
    {"od": 108.0, "wall": 2.5, "id": 103.0, "label": "108 × 2,5"},
]

# ─────────────────────────────────────────────────────────────────────────────
# KÄLTEMITTEL-DATENTABELLEN  (Sättigungseigenschaften)
# Spalten: temps [°C], p_bar, rho_v [kg/m³], rho_l [kg/m³],
#          h_fg [kJ/kg], mu_v [μPa·s]
# ─────────────────────────────────────────────────────────────────────────────
REFRIGERANTS = {
    "R744": {
        "name": "R744 (CO₂)", "color": "#FF6B35",
        "warning": "CO₂-Hochdruckanlage! K65-Armaturen erforderlich. Sicherheitsventile einplanen. Max. Betriebsdruck beachten.",
        "a2l": False, "co2": True,
        "temps":  [-50,   -40,   -35,   -30,   -25,   -20,   -15,   -10,   0,     10,    20,    30],
        "p_bar":  [6.83,  10.05, 12.05, 14.34, 16.96, 19.94, 23.31, 27.12, 34.85, 45.01, 57.34, 72.13],
        "rho_v":  [14.5,  21.3,  25.9,  31.2,  37.5,  44.9,  53.7,  64.1,  92.0,  134.0, 202.0, 316.0],
        "rho_l":  [1153,  1119,  1101,  1082,  1062,  1039,  1015,  989,   929,   858,   771,   655],
        "h_fg":   [322,   312,   306,   299,   292,   283,   273,   261,   231,   193,   143,   72],
        "mu_v":   [11.0,  11.5,  11.8,  12.1,  12.4,  12.8,  13.2,  13.6,  14.7,  16.1,  18.0,  21.5],
    },
    "R1234yf": {
        "name": "R1234yf", "color": "#4CAF50",
        "warning": "A2L-Kältemittel — Zündschutz nach EN 378 / EN 60335-2-40 erforderlich.",
        "a2l": True, "co2": False,
        "temps":  [-50,  -40,  -35,  -30,  -25,  -20,  -15,  -10,  0,    10,   20,   30,   40,   50],
        "p_bar":  [0.51, 0.80, 1.00, 1.25, 1.55, 1.91, 2.34, 2.84, 4.07, 5.72, 7.86, 10.57,13.97,18.17],
        "rho_v":  [2.8,  4.4,  5.5,  6.9,  8.6,  10.6, 13.1, 16.1, 23.1, 33.2, 46.9, 65.7, 91.7, 128.0],
        "rho_l":  [1237, 1200, 1181, 1162, 1143, 1123, 1102, 1081, 1036, 988,  934,  874,  803,  716],
        "h_fg":   [211,  206,  203,  200,  197,  194,  190,  186,  178,  168,  155,  139,  119,  93],
        "mu_v":   [8.5,  9.0,  9.2,  9.5,  9.8,  10.1, 10.4, 10.8, 11.5, 12.3, 13.3, 14.6, 16.2, 18.5],
    },
    "R455A": {
        "name": "R455A", "color": "#9C27B0",
        "warning": "A2L-Kältemittel (R744/R134a/R1234yf) — Zündschutz beachten.",
        "a2l": True, "co2": False,
        "temps":  [-50,  -40,  -35,  -30,  -25,  -20,  -15,  -10,  0,    10,   20,   30,   40,   50],
        "p_bar":  [0.75, 1.15, 1.45, 1.80, 2.20, 2.70, 3.25, 3.90, 5.45, 7.50, 10.10,13.35,17.30,22.10],
        "rho_v":  [4.5,  6.9,  8.7,  10.8, 13.4, 16.5, 20.2, 24.5, 34.8, 48.8, 67.5, 92.0, 124.0,166.0],
        "rho_l":  [1220, 1185, 1167, 1148, 1129, 1109, 1088, 1067, 1022, 972,  915,  850,  773,  680],
        "h_fg":   [300,  295,  290,  285,  280,  274,  268,  261,  247,  230,  210,  186,  158,  122],
        "mu_v":   [9.0,  9.4,  9.7,  10.0, 10.3, 10.7, 11.1, 11.5, 12.4, 13.4, 14.6, 16.2, 18.2, 21.0],
    },
    "R452A": {
        "name": "R452A", "color": "#FF9800",
        "warning": "A2L-Kältemittel (R32/R125/R1234yf) — Zündschutz beachten.",
        "a2l": True, "co2": False,
        "temps":  [-50,  -40,  -35,  -30,  -25,  -20,  -15,  -10,  0,    10,   20,   30,   40,   50],
        "p_bar":  [1.05, 1.60, 1.97, 2.40, 2.90, 3.50, 4.20, 5.00, 6.90, 9.30, 12.30,16.00,20.50,25.90],
        "rho_v":  [6.2,  9.4,  11.7, 14.5, 17.8, 21.8, 26.5, 32.0, 45.0, 62.5, 85.5, 115.0,154.0,205.0],
        "rho_l":  [1330, 1295, 1277, 1258, 1238, 1218, 1197, 1175, 1129, 1079, 1023, 958,  882,  790],
        "h_fg":   [238,  233,  230,  226,  222,  218,  213,  208,  197,  184,  168,  149,  125,  96],
        "mu_v":   [9.5,  9.9,  10.2, 10.6, 11.0, 11.4, 11.8, 12.3, 13.3, 14.5, 15.9, 17.8, 20.2, 23.5],
    },
    "R513A": {
        "name": "R513A", "color": "#00BCD4",
        "warning": "HFO-Blend (R1234yf/R134a) — Sicherheitsklasse A1.",
        "a2l": False, "co2": False,
        "temps":  [-50,  -40,  -35,  -30,  -25,  -20,  -15,  -10,  0,    10,   20,   30,   40,   50],
        "p_bar":  [0.64, 0.99, 1.24, 1.54, 1.89, 2.32, 2.82, 3.41, 4.88, 6.80, 9.26, 12.34,16.15,20.80],
        "rho_v":  [3.5,  5.5,  6.9,  8.6,  10.7, 13.2, 16.3, 19.9, 28.8, 41.0, 57.5, 79.5, 109.0,149.0],
        "rho_l":  [1288, 1254, 1237, 1219, 1201, 1182, 1162, 1141, 1097, 1049, 995,  933,  860,  770],
        "h_fg":   [208,  203,  200,  197,  193,  189,  185,  181,  172,  161,  148,  131,  110,  83],
        "mu_v":   [9.0,  9.4,  9.7,  10.0, 10.3, 10.7, 11.1, 11.5, 12.4, 13.4, 14.7, 16.3, 18.5, 21.5],
    },
    "R32": {
        "name": "R32 (Samsung EHS/Quint)", "color": "#1565C0",
        "warning": "A2L-Kältemittel! Zündschutz nach EN 378-1 / EN 60335-2-40 erforderlich. GWP = 675. Typisch in Samsung EHS Quint, Split- und VRF-Anlagen.",
        "a2l": True, "co2": False,
        "temps":  [-50,   -40,   -35,   -30,   -25,   -20,   -15,   -10,   0,     10,    20,    30,    40,    50],
        "p_bar":  [1.69,  2.56,  3.12,  3.73,  4.50,  5.27,  6.22,  7.30,  9.95,  13.22, 17.19, 21.96, 27.60, 34.31],
        "rho_v":  [9.0,   13.5,  16.5,  19.7,  23.5,  27.9,  33.0,  38.9,  53.5,  72.5,  97.3,  129.0, 170.0, 224.0],
        "rho_l":  [1199,  1164,  1146,  1128,  1108,  1089,  1069,  1047,  1000,  949,   893,   831,   760,   678],
        "h_fg":   [392,   380,   374,   367,   360,   352,   344,   335,   316,   294,   269,   239,   203,   158],
        "mu_v":   [10.5,  10.9,  11.1,  11.4,  11.7,  11.9,  12.2,  12.5,  13.2,  14.0,  15.0,  16.3,  18.0,  20.5],
    },
    "R449A": {
        "name": "R449A", "color": "#E91E63",
        "warning": "HFO-Blend (R32/R125/R1234yf/R134a) — Sicherheitsklasse A1.",
        "a2l": False, "co2": False,
        "temps":  [-50,  -40,  -35,  -30,  -25,  -20,  -15,  -10,  0,    10,   20,   30,   40,   50],
        "p_bar":  [1.10, 1.68, 2.07, 2.52, 3.05, 3.66, 4.37, 5.20, 7.20, 9.70, 12.85,16.75,21.50,27.20],
        "rho_v":  [6.5,  9.9,  12.3, 15.3, 18.9, 23.2, 28.2, 34.2, 48.5, 67.5, 92.5, 125.0,168.0,224.0],
        "rho_l":  [1315, 1280, 1262, 1243, 1224, 1203, 1182, 1160, 1113, 1060, 1000, 932,  852,  756],
        "h_fg":   [260,  255,  252,  248,  243,  238,  232,  225,  211,  194,  175,  152,  124,  90],
        "mu_v":   [9.8,  10.2, 10.6, 11.0, 11.4, 11.8, 12.3, 12.8, 13.9, 15.2, 16.8, 18.8, 21.5, 25.0],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# PHYSIKFUNKTIONEN
# ─────────────────────────────────────────────────────────────────────────────
def interp_prop(ref_key, T_C, prop):
    d = REFRIGERANTS[ref_key]
    return float(np.interp(T_C, d["temps"], d[prop]))

def get_sat_props(ref_key, T_C):
    return {
        "p_bar":  interp_prop(ref_key, T_C, "p_bar"),
        "rho_v":  interp_prop(ref_key, T_C, "rho_v"),
        "rho_l":  interp_prop(ref_key, T_C, "rho_l"),
        "h_fg":   interp_prop(ref_key, T_C, "h_fg"),
        "mu_v":   interp_prop(ref_key, T_C, "mu_v") * 1e-6,  # μPa·s → Pa·s
    }

def darcy_f(Re, eps_m=1.5e-6):
    """Darcy-Weisbach Reibungszahl für glattes CU-Rohr (ε≈0.0015 mm)."""
    if Re < 1:
        return 1.0
    if Re < 2300:
        return 64.0 / Re
    if Re < 4000:
        return 0.316 / (Re ** 0.25)
    # Colebrook-White (3 Iterationen)
    d = 0.025  # Schätzwert für d – wird als relatives Rauheitsmaß intern nur für f benötigt
    f = 0.316 / (Re ** 0.25)
    return f  # Für glatte Cu-Rohre reicht Blasius bis Re~1e5; darüber Churchill:
    # f_inv = -2*log10(eps/3.7 + 2.51/(Re*sqrt(f))); iterieren

def calc_pipe(pipe, m_dot_kg_s, rho_kg_m3, mu_Pa_s, L_eq_m, T_sat_C, h_fg_kJ, rho_v_kg_m3):
    """Berechne v, Δp für ein Rohr.  Gibt dict zurück."""
    d_m = pipe["id"] / 1000.0
    A   = math.pi * (d_m / 2.0) ** 2
    v   = m_dot_kg_s / (rho_kg_m3 * A) if A > 0 else 999.0
    Re  = rho_kg_m3 * abs(v) * d_m / mu_Pa_s if mu_Pa_s > 0 else 1e6
    f   = darcy_f(Re)
    dp_Pa  = f * (L_eq_m / d_m) * rho_kg_m3 * v**2 / 2.0
    dp_bar = dp_Pa / 1e5
    # Clausius-Clapeyron: ΔT = ΔP / (h_fg * ρv / T_sat)
    T_sat_K = T_sat_C + 273.15
    dp_dT   = (h_fg_kJ * 1000.0 * rho_v_kg_m3) / T_sat_K  # Pa/K
    dp_K    = dp_Pa / dp_dT if dp_dT > 0 else 0.0
    return {"v": v, "Re": Re, "dp_Pa": dp_Pa, "dp_bar": dp_bar, "dp_K": dp_K}

def select_pipe(m_dot_kg_s, rho_kg_m3, mu_Pa_s, L_eq_m,
                v_min, v_max, dp_K_max, T_sat_C, h_fg_kJ, rho_v_kg_m3):
    """
    Wählt optimalen Rohrdurchmesser.
    Gibt (auto_idx, warnings) zurück.
    Priorität:
      1. v_min ≤ v ≤ v_max  UND  Δp ≤ Grenzwert  → kleinstes solches
      2. v ≤ v_max  UND  Δp ≤ Grenzwert  → kleinstes (v_min-Warnung)
      3. v_min ≤ v ≤ v_max  UND  Δp > Grenzwert  → größtes solches (Δp-Warnung)
      4. Fallback → größtes verfügbares Rohr
    """
    best = {"prio": 99, "idx": len(CU_PIPES) - 1, "warns": ["⚠️ Kein ideales Rohr gefunden — größtes Rohr gewählt"]}

    for i, pipe in enumerate(CU_PIPES):
        r = calc_pipe(pipe, m_dot_kg_s, rho_kg_m3, mu_Pa_s, L_eq_m, T_sat_C, h_fg_kJ, rho_v_kg_m3)
        v_ok  = v_min <= r["v"] <= v_max
        dp_ok = r["dp_K"] <= dp_K_max
        v_max_ok = r["v"] <= v_max
        v_min_ok = r["v"] >= v_min

        if v_ok and dp_ok:
            return i, []  # perfekt → sofort zurück
        elif v_max_ok and dp_ok and best["prio"] > 2:
            best = {"prio": 2, "idx": i,
                    "warns": [f"⚠️ v = {r['v']:.1f} m/s < v_min {v_min:.1f} m/s — Ölrückführung prüfen! (DSR empfohlen)"]}
        elif v_ok and not dp_ok and best["prio"] > 3:
            best = {"prio": 3, "idx": i,
                    "warns": [f"⚠️ Δp = {r['dp_K']:.2f} K > Grenzwert {dp_K_max:.1f} K — Rohr zu klein"]}
        elif v_max_ok and best["prio"] > 4:
            best = {"prio": 4, "idx": i,
                    "warns": [f"⚠️ v = {r['v']:.1f} m/s < v_min, Δp = {r['dp_K']:.2f} K — Kompromiss"]}

    return best["idx"], best["warns"]

def dp_limit_K(line_type, L_total_m, app_type):
    """Längenabhängiger Druckverlustgrenzwert nach Technik-Bibel."""
    if app_type == "NK":
        if line_type == "SL":
            if L_total_m <= 25:  return 1.5
            elif L_total_m <= 50: return 1.2
            else:                 return 1.0
        elif line_type == "DL":   return 2.0
        elif line_type == "FL":   return 0.5
        else:                     return 1.5
    else:  # TK
        if line_type == "SL":
            if L_total_m <= 25:  return 1.0
            elif L_total_m <= 50: return 0.8
            else:                 return 0.6
        elif line_type == "DL":   return 1.5
        elif line_type == "FL":   return 0.4
        else:                     return 1.0

def equiv_length(L_m, n_elbows, n_ball_valves, n_solenoid, d_od_mm):
    """Äquivalente Leitungslänge inkl. Formstücke (nach VDI 2067 / Kältetechnik)."""
    d_m = d_od_mm / 1000.0
    L_eq = L_m
    L_eq += n_elbows     * 1.2 * d_m * 30   # 90°-Bogen ≈ 30×d
    L_eq += n_ball_valves * 1.2 * d_m * 6    # Kugelhahn ≈ 6×d
    L_eq += n_solenoid   * 1.2 * d_m * 75   # Magnetventil ≈ 75×d
    return L_eq

def hydrostatic_dp(rho, h_m):
    """Hydrostatischer Druckunterschied in Pa (h>0 = Steigung)."""
    return rho * 9.81 * h_m

def dew_point_C(T_amb_C, phi_pct):
    """Magnus-Formel Taupunkt."""
    a, b = 17.625, 243.04
    phi = phi_pct / 100.0
    if phi <= 0: return -100.0
    gamma = (a * T_amb_C / (b + T_amb_C)) + math.log(phi)
    return b * gamma / (a - gamma)

def insulation_thickness_mm(T_pipe_C, T_amb_C, phi_pct, d_od_mm):
    """Mindest-Dämmdicke (Armaflex) nach Taupunktbedingung."""
    T_dew = dew_point_C(T_amb_C, phi_pct)
    if T_pipe_C >= T_dew:
        return 0  # keine Dämmung nötig
    dT = T_dew - T_pipe_C
    # Faustformel: λ_Armaflex ≈ 0.038 W/mK, U-Wert-Betrachtung
    # δ = d_od/2 * (exp(λ * ln(2) / (α * dT)) - 1)  vereinfacht:
    alpha_o = 10.0  # W/m²K Außenübergang
    lam = 0.038     # W/mK
    r_pipe = d_od_mm / 2000.0  # m
    delta_m = r_pipe * (math.exp(lam * math.log(2) / (alpha_o * r_pipe * max(dT, 0.1) / 10.0)) - 1)
    # Praktisch: Standardstufen
    raw = max(dT * 2.5, 9.0)  # mm
    for std in [9, 13, 19, 25, 32, 40]:
        if std >= raw:
            return std
    return 40

# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT SETUP & CSS
# ─────────────────────────────────────────────────────────────────────────────
try:
    st.set_page_config(page_title="°coolROHR", layout="wide", page_icon="🔧")
except Exception:
    pass

COLOR_BLUE  = "#36A9E1"
COLOR_GRAY  = "#3C3C3B"
COLOR_WHITE = "#FFFFFF"
COLOR_GREEN = "#4CAF50"
COLOR_RED   = "#f44336"
COLOR_AMBER = "#FF9800"

def get_font_b64():
    for p in ["POE Vetica UI.ttf", "POE_Vetica_UI.ttf"]:
        if os.path.exists(p):
            with open(p, "rb") as f:
                return base64.b64encode(f.read()).decode()
    return None

def get_logo_b64():
    for p in ["Coolsulting_Logo_ohneHG_blau.png", "Coolsulting_Logo_ohneHG_blau_weiß.png"]:
        if os.path.exists(p):
            with open(p, "rb") as f:
                return base64.b64encode(f.read()).decode()
    return None

font_b64 = get_font_b64()
logo_b64 = get_logo_b64()

st.markdown(f"""
<style>
{f"@font-face{{font-family:'POE Helvetica UI';src:url(data:font/ttf;base64,{font_b64}) format('truetype');}}" if font_b64 else ""}
*{{font-family:'POE Helvetica UI','Helvetica Neue',Helvetica,sans-serif!important;}}
.stApp{{background:{COLOR_BLUE};}}
section[data-testid="stSidebar"]{{background:{COLOR_GRAY}!important;}}
section[data-testid="stSidebar"] *{{color:white!important;}}
section[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"]>div{{background:rgba(255,255,255,0.12)!important;color:white!important;}}
section[data-testid="stSidebar"] input{{background:rgba(255,255,255,0.12)!important;color:white!important;}}
.main-header{{display:flex;justify-content:space-between;align-items:flex-end;margin-bottom:18px;}}
.title-big{{font-size:52px;font-weight:bold;line-height:1;}}
.cool-part{{color:white;}}
.rohr-part{{color:{COLOR_GRAY};text-transform:uppercase;}}
.sub-right{{color:rgba(255,255,255,0.8);font-size:14px;text-align:right;padding-bottom:4px;}}
.card{{background:white;border-radius:12px;padding:20px;margin-bottom:14px;}}
.card-title{{color:{COLOR_BLUE};font-weight:bold;font-size:16px;border-bottom:2px solid {COLOR_BLUE};padding-bottom:6px;margin-bottom:12px;}}
.metric-box{{background:{COLOR_GRAY};border-radius:8px;padding:12px 16px;text-align:center;color:white;}}
.metric-val{{font-size:24px;font-weight:bold;}}
.metric-lbl{{font-size:11px;opacity:0.8;margin-top:2px;}}
.status-ok{{color:{COLOR_GREEN};font-weight:bold;}}
.status-warn{{color:{COLOR_AMBER};font-weight:bold;}}
.status-err{{color:{COLOR_RED};font-weight:bold;}}
.warn-box{{background:#FFF3E0;border-left:4px solid {COLOR_AMBER};border-radius:6px;padding:10px 14px;margin:8px 0;color:#7f4f00;font-size:13px;}}
.err-box{{background:#FFEBEE;border-left:4px solid {COLOR_RED};border-radius:6px;padding:10px 14px;margin:8px 0;color:#b71c1c;font-size:13px;}}
.info-box{{background:#E3F2FD;border-left:4px solid {COLOR_BLUE};border-radius:6px;padding:10px 14px;margin:8px 0;color:#0d47a1;font-size:13px;}}
.pipe-label{{font-size:20px;font-weight:bold;color:{COLOR_GRAY};}}
.v-bar-outer{{height:18px;background:#e0e0e0;border-radius:9px;overflow:hidden;margin:4px 0;}}
.result-table{{width:100%;border-collapse:collapse;font-size:13px;}}
.result-table th{{background:{COLOR_GRAY};color:white;padding:7px 10px;text-align:left;}}
.result-table td{{padding:6px 10px;border-bottom:1px solid #eee;}}
.result-table tr:last-child td{{border-bottom:none;}}
.dsr-badge{{display:inline-block;background:{COLOR_BLUE};color:white;border-radius:12px;padding:2px 10px;font-size:12px;font-weight:bold;margin-left:8px;}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:48px;">' if logo_b64 else ""
st.markdown(f"""
<div class="main-header">
  <div class="title-big"><span class="cool-part">°cool</span><span class="rohr-part">ROHR</span></div>
  <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px;">
    {logo_html}
    <div class="sub-right">Kältemittel-Rohrdimensionierung | v2.0 | {datetime.now().strftime('%d.%m.%Y')}</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "calculated": False,
        "offsets": {"SL": 0, "DL": 0, "FL": 0, "KL": 0,
                    "SL_speed": 0, "SL_main": 0, "DL_speed": 0, "DL_main": 0},
        "dsr_sl": False,
        "dsr_dl": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR – EINGABEN
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"<div style='font-size:18px;font-weight:bold;color:{COLOR_BLUE};border-bottom:2px solid {COLOR_BLUE};padding-bottom:6px;margin-bottom:14px;'>PROJEKTEINGABEN</div>", unsafe_allow_html=True)

    kunde     = st.text_input("Kunde", "Musterkunde GmbH")
    projekt   = st.text_input("Projekt", "Kälteanlage NK")
    bearbeiter= st.text_input("Bearbeiter", "M. Schäpers")

    st.markdown("---")
    st.markdown("**🌡️ Anlage**")
    ref_key   = st.selectbox("Kältemittel", list(REFRIGERANTS.keys()),
                              format_func=lambda k: REFRIGERANTS[k]["name"])
    app_type  = st.radio("Anwendung", ["NK (Normalkühlung)", "TK (Tiefkühlung)"],
                         horizontal=True)
    app_code  = "NK" if app_type.startswith("NK") else "TK"

    t0_default = -10.0 if app_code == "NK" else -35.0
    t0  = st.number_input("Verdampfungstemperatur t₀ (°C)", -50.0, 15.0, t0_default, 1.0)
    tc  = st.number_input("Verflüssigungstemperatur tc (°C)", 20.0, 70.0, 40.0, 1.0)
    Q_kW= st.number_input("Kälteleistung Q₀ (kW)", 0.5, 5000.0, 10.0, 0.5)

    st.markdown("---")
    st.markdown("**📏 Leitungslängen (m)**")
    L_SL_h= st.number_input("Saugleitung horizontal", 0.0, 500.0, 10.0, 1.0)
    L_SL_v= st.number_input("Saugleitung vertikal (Steigleitung)", 0.0, 200.0, 3.0, 0.5)
    L_DL  = st.number_input("Druckleitung gesamt", 0.0, 200.0, 0.5, 0.5)
    same_FL= st.checkbox("Flüssigkeitsleitung = Saugleitung", True)
    L_FL  = L_SL_h + L_SL_v if same_FL else st.number_input("Flüssigkeitsleitung", 0.0, 500.0, 10.0, 1.0)
    L_KL  = st.number_input("Kondensatleitung (Flüssigkeit nach Kondensator)", 0.0, 100.0, 0.5, 0.5)

    st.markdown("---")
    st.markdown("**📐 Formstücke (je Leitung)**")
    col_e, col_b, col_s = st.columns(3)
    n_el = col_e.number_input("Bögen", 0, 30, 4, 1)
    n_bv = col_b.number_input("Kugelhähne", 0, 20, 1, 1)
    n_sv = col_s.number_input("Magnetventile", 0, 10, 1, 1)

    st.markdown("---")
    st.markdown("**↕️ Höhenunterschiede**")
    h_SL  = st.number_input("Steigleitung Saugleitung (m, +=nach oben)", -30.0, 30.0, 3.0, 0.5)
    h_FL  = st.number_input("Höhe Flüssigkeitsleitung (m, –=nach unten=Gewinn)", -30.0, 30.0, -3.0, 0.5)

    st.markdown("---")
    st.markdown("**🌍 Umgebung (für Isolierung)**")
    T_amb = st.number_input("Umgebungstemperatur (°C)", 10.0, 45.0, 25.0, 1.0)
    phi   = st.number_input("Relative Feuchte (%)", 30.0, 100.0, 70.0, 5.0)

    st.markdown("---")
    calc_btn = st.button("⚡ BERECHNEN", use_container_width=True,
                         type="primary")

if calc_btn:
    # Offsets bei neuer Berechnung zurücksetzen
    st.session_state.offsets = {k: 0 for k in st.session_state.offsets}
    st.session_state.dsr_sl = False
    st.session_state.dsr_dl = False
    st.session_state.calculated = True

# ─────────────────────────────────────────────────────────────────────────────
# HILFSFUNKTIONEN FÜR DIE ERGEBNISDARSTELLUNG
# ─────────────────────────────────────────────────────────────────────────────
def velocity_bar_html(v, v_min, v_max, label=""):
    """Farbiger Geschwindigkeitsbalken mit Markierungen."""
    v_scale_max = max(v_max * 1.3, v * 1.1)
    pct_v    = min(v / v_scale_max * 100, 100)
    pct_vmin = v_min / v_scale_max * 100
    pct_vmax = v_max / v_scale_max * 100
    if v < v_min:      bar_color = COLOR_RED
    elif v > v_max:    bar_color = COLOR_RED
    else:              bar_color = COLOR_GREEN
    status_txt = "✅ OK" if v_min <= v <= v_max else ("⬇️ zu niedrig" if v < v_min else "⬆️ zu hoch")
    return f"""
<div style='margin:8px 0;'>
  <div style='display:flex;justify-content:space-between;font-size:12px;margin-bottom:2px;'>
    <span>{label}</span><span style='font-weight:bold;color:{bar_color};'>{v:.2f} m/s — {status_txt}</span>
  </div>
  <div style='position:relative;height:18px;background:#e0e0e0;border-radius:9px;overflow:visible;'>
    <div style='position:absolute;left:0;top:0;height:100%;width:{pct_v:.1f}%;background:{bar_color};border-radius:9px;'></div>
    <div style='position:absolute;left:{pct_vmin:.1f}%;top:-4px;height:26px;width:2px;background:#1565C0;'></div>
    <div style='position:absolute;left:{pct_vmax:.1f}%;top:-4px;height:26px;width:2px;background:#b71c1c;'></div>
  </div>
  <div style='display:flex;justify-content:space-between;font-size:10px;margin-top:2px;color:#666;'>
    <span>v_min {v_min:.1f} m/s</span><span>v_max {v_max:.1f} m/s</span>
  </div>
</div>"""

def dp_bar_html(dp_K, dp_K_max, dp_bar):
    pct = min(dp_K / dp_K_max * 100, 130)
    color = COLOR_GREEN if dp_K <= dp_K_max else COLOR_RED
    return f"""
<div style='margin:6px 0;'>
  <div style='display:flex;justify-content:space-between;font-size:12px;margin-bottom:2px;'>
    <span>Druckverlust</span>
    <span style='font-weight:bold;color:{color};'>Δp = {dp_K:.3f} K | {dp_bar*1000:.1f} mbar</span>
  </div>
  <div style='position:relative;height:14px;background:#e0e0e0;border-radius:7px;overflow:hidden;'>
    <div style='height:100%;width:{min(pct,100):.1f}%;background:{color};border-radius:7px;'></div>
  </div>
  <div style='font-size:10px;color:#666;margin-top:2px;'>Grenzwert: {dp_K_max:.1f} K</div>
</div>"""

def pipe_selector_widget(key, auto_idx, label_prefix):
    """Plus/Minus-Buttons für manuellen Rohrgröße-Override. Gibt aktuellen Index zurück."""
    offset = st.session_state.offsets[key]
    cur_idx = max(0, min(len(CU_PIPES) - 1, auto_idx + offset))
    c1, c2, c3 = st.columns([1, 4, 1])
    with c1:
        if st.button("−", key=f"btn_minus_{key}", help="Kleineres Rohr"):
            st.session_state.offsets[key] = max(-auto_idx, offset - 1)
            st.rerun()
    with c2:
        pipe = CU_PIPES[cur_idx]
        mode = "✎ MANUELL" if offset != 0 else "AUTO"
        delta_txt = f" ({'▲' if offset > 0 else '▼'}{abs(offset)})" if offset != 0 else ""
        st.markdown(f"<div style='text-align:center;padding:4px 0;'>"
                    f"<span style='font-size:20px;font-weight:bold;color:{COLOR_GRAY};'>⌀ {pipe['od']:.0f} / {pipe['id']:.0f} mm</span><br>"
                    f"<span style='font-size:12px;color:#888;'>{pipe['label']} mm | {mode}{delta_txt}</span></div>",
                    unsafe_allow_html=True)
    with c3:
        if st.button("+", key=f"btn_plus_{key}", help="Größeres Rohr"):
            st.session_state.offsets[key] = min(len(CU_PIPES) - 1 - auto_idx, offset + 1)
            st.rerun()
    return cur_idx

def show_pipe_result(title, line_code, auto_idx, m_dot, rho, mu, L_eq, v_min, v_max,
                     dp_K_max, T_sat_C, h_fg, rho_v, dp_hydro_Pa=0.0, warns=[]):
    """Zeigt Ergebnisblock für eine Leitung mit manueller Übersteuerung."""
    st.markdown(f"<div class='card-title'>{title}</div>", unsafe_allow_html=True)
    cur_idx = pipe_selector_widget(line_code, auto_idx, title)
    pipe = CU_PIPES[cur_idx]
    r = calc_pipe(pipe, m_dot, rho, mu, L_eq, T_sat_C, h_fg, rho_v)

    # Hydrostatik-Korrektur
    dp_total_Pa = r["dp_Pa"] + dp_hydro_Pa
    dp_bar_total = dp_total_Pa / 1e5
    T_sat_K = T_sat_C + 273.15
    dp_dT = (h_fg * 1000.0 * rho_v) / T_sat_K
    dp_K_total = dp_total_Pa / dp_dT if dp_dT > 0 else r["dp_K"]

    st.markdown(velocity_bar_html(r["v"], v_min, v_max, "Strömungsgeschwindigkeit"), unsafe_allow_html=True)
    st.markdown(dp_bar_html(dp_K_total, dp_K_max, dp_bar_total), unsafe_allow_html=True)

    # Kennwerte-Tabelle
    insul = insulation_thickness_mm(T_sat_C, T_amb, phi, pipe["od"])
    hydro_txt = f"{dp_hydro_Pa/1e5*1000:.1f} mbar ({'Verlust' if dp_hydro_Pa > 0 else 'Gewinn'})" if dp_hydro_Pa != 0 else "—"
    st.markdown(f"""
<table class='result-table'>
  <tr><td>Massenstrom ṁ</td><td><b>{m_dot*3600:.1f} kg/h  ({m_dot:.4f} kg/s)</b></td></tr>
  <tr><td>Dichte ρ</td><td><b>{rho:.2f} kg/m³</b></td></tr>
  <tr><td>Reibungsdruckverlust</td><td><b>{r['dp_K']:.3f} K  |  {r['dp_bar']*1000:.1f} mbar</b></td></tr>
  <tr><td>Hydrostatik</td><td><b>{hydro_txt}</b></td></tr>
  <tr><td>Gesamt-Δp</td><td><b>{dp_K_total:.3f} K  |  {dp_bar_total*1000:.1f} mbar</b></td></tr>
  <tr><td>Re-Zahl</td><td><b>{r['Re']:,.0f}</b></td></tr>
  <tr><td>Isolierung (Armaflex)</td><td><b>{insul} mm</b></td></tr>
</table>""", unsafe_allow_html=True)

    for w in warns:
        cls = "err-box" if "Ölrückführung" in w or "DSR" in w else "warn-box"
        st.markdown(f"<div class='{cls}'>{w}</div>", unsafe_allow_html=True)

    return r, dp_K_total

# ─────────────────────────────────────────────────────────────────────────────
# SVG-SCHEMATA
# ─────────────────────────────────────────────────────────────────────────────
def show_dsr_schema():
    """Zeigt 4 SVG-Schaltungsschemas für DSR."""

    svg_css = "style='width:100%;max-width:380px;display:block;margin:auto;'"

    # Schema 1: Gesamtübersicht
    svg1 = f"""<svg {svg_css} viewBox="0 0 300 420" xmlns="http://www.w3.org/2000/svg">
  <rect width="300" height="420" fill="#f8f9fa" rx="8"/>
  <text x="150" y="22" text-anchor="middle" font-size="13" font-weight="bold" fill="{COLOR_GRAY}">1 — Gesamtübersicht DSR</text>
  <!-- Verdampfer unten -->
  <rect x="100" y="360" width="100" height="40" fill="{COLOR_BLUE}" rx="4"/>
  <text x="150" y="384" text-anchor="middle" font-size="11" fill="white">Verdampfer</text>
  <!-- Verdichter oben -->
  <rect x="100" y="30" width="100" height="36" fill="{COLOR_GRAY}" rx="4"/>
  <text x="150" y="52" text-anchor="middle" font-size="11" fill="white">Verdichter</text>
  <!-- Speed Riser (kleines Rohr, links) -->
  <line x1="120" y1="360" x2="120" y2="180" stroke="{COLOR_BLUE}" stroke-width="3"/>
  <text x="106" y="280" text-anchor="middle" font-size="9" fill="{COLOR_BLUE}" transform="rotate(-90,106,280)">Speed Riser</text>
  <!-- Ölfallen am Speed Riser (alle 4m) -->
  <path d="M 120 260 Q 108 260 108 270 Q 108 280 120 280" fill="none" stroke="{COLOR_BLUE}" stroke-width="2"/>
  <ellipse cx="108" cy="275" rx="4" ry="3" fill="#8B4513" opacity="0.7"/>
  <text x="92" y="278" text-anchor="middle" font-size="7" fill="#8B4513">Öl</text>
  <path d="M 120 200 Q 108 200 108 210 Q 108 220 120 220" fill="none" stroke="{COLOR_BLUE}" stroke-width="2"/>
  <ellipse cx="108" cy="215" rx="4" ry="3" fill="#8B4513" opacity="0.7"/>
  <!-- Main Riser (großes Rohr, rechts) -->
  <line x1="180" y1="360" x2="180" y2="180" stroke="{COLOR_RED}" stroke-width="5"/>
  <text x="197" y="280" text-anchor="middle" font-size="9" fill="{COLOR_RED}" transform="rotate(90,197,280)">Main Riser</text>
  <!-- U-Bogen (Ölfalle) am Main Riser Fuß -->
  <path d="M 180 360 Q 180 376 168 376 Q 156 376 156 360" fill="none" stroke="{COLOR_RED}" stroke-width="5"/>
  <ellipse cx="168" cy="376" rx="6" ry="4" fill="#8B4513" opacity="0.8"/>
  <text x="168" y="393" text-anchor="middle" font-size="8" fill="#8B4513">Ölfalle</text>
  <!-- Abreisbögen oben (a + b) -->
  <!-- Speed Riser Abreisbogen (a) -->
  <path d="M 120 180 L 120 168 Q 120 155 132 155 Q 144 155 144 168 L 144 178 L 152 178" fill="none" stroke="{COLOR_BLUE}" stroke-width="3"/>
  <text x="122" y="150" font-size="8" fill="{COLOR_BLUE}">(a)</text>
  <!-- Main Riser Abreisbogen (b) -->
  <path d="M 180 180 L 180 168 Q 180 155 168 155 Q 156 155 156 168 L 156 178 L 152 178" fill="none" stroke="{COLOR_RED}" stroke-width="5"/>
  <text x="182" y="150" font-size="8" fill="{COLOR_RED}">(b)</text>
  <!-- T-Stück -->
  <rect x="146" y="174" width="12" height="8" fill="{COLOR_GRAY}" rx="2"/>
  <line x1="152" y1="174" x2="152" y2="66" stroke="{COLOR_GRAY}" stroke-width="4"/>
  <!-- Legende -->
  <circle cx="22" cy="200" r="5" fill="{COLOR_BLUE}"/><text x="32" y="204" font-size="9" fill="{COLOR_GRAY}">Speed Riser</text>
  <rect x="17" y="215" width="10" height="6" fill="{COLOR_RED}"/><text x="32" y="222" font-size="9" fill="{COLOR_GRAY}">Main Riser</text>
  <ellipse cx="22" cy="238" rx="5" ry="3" fill="#8B4513"/><text x="32" y="241" font-size="9" fill="{COLOR_GRAY}">Öl-Sohle</text>
  <!-- Pfeile Strömung -->
  <text x="92" y="358" font-size="10" fill="{COLOR_BLUE}">↑</text>
  <text x="183" y="358" font-size="10" fill="{COLOR_RED}">↑</text>
</svg>"""

    # Schema 2: U-Bogen Ölfalle Detail
    svg2 = f"""<svg {svg_css} viewBox="0 0 300 320" xmlns="http://www.w3.org/2000/svg">
  <rect width="300" height="320" fill="#f8f9fa" rx="8"/>
  <text x="150" y="22" text-anchor="middle" font-size="13" font-weight="bold" fill="{COLOR_GRAY}">2 — U-Bogen Ölfalle (Fuß Main Riser)</text>
  <!-- Main Riser kommt von unten -->
  <line x1="150" y1="300" x2="150" y2="200" stroke="{COLOR_RED}" stroke-width="8"/>
  <!-- U-Bogen -->
  <path d="M 150 200 Q 150 170 130 170 Q 110 170 110 200 L 110 230" fill="none" stroke="{COLOR_RED}" stroke-width="8"/>
  <!-- Rechter Arm hoch -->
  <line x1="110" y1="230" x2="110" y2="130" stroke="{COLOR_RED}" stroke-width="8"/>
  <!-- Öl-Sohle -->
  <ellipse cx="130" cy="175" rx="18" ry="8" fill="#8B4513" opacity="0.7"/>
  <text x="130" y="178" text-anchor="middle" font-size="10" fill="white">Öl-Sohle</text>
  <!-- Beschriftung -->
  <text x="160" y="260" font-size="11" fill="{COLOR_RED}">↑ vom Verdampfer</text>
  <text x="55" y="125" font-size="11" fill="{COLOR_RED}">↑ zum Verdichter</text>
  <!-- Erklärung -->
  <rect x="10" y="40" width="280" height="60" fill="white" rx="6" stroke="{COLOR_BLUE}" stroke-width="1"/>
  <text x="20" y="60" font-size="10" fill="{COLOR_GRAY}">Bei Teillast: Öl sammelt sich in der Sohle</text>
  <text x="20" y="75" font-size="10" fill="{COLOR_GRAY}">→ U-Bogen hydraulisch gesperrt</text>
  <text x="20" y="90" font-size="10" fill="{COLOR_GRAY}">→ Strom fließt durch Speed Riser</text>
</svg>"""

    # Schema 3: Gooseneck / Abreisbogen Detail
    svg3 = f"""<svg {svg_css} viewBox="0 0 300 360" xmlns="http://www.w3.org/2000/svg">
  <rect width="300" height="360" fill="#f8f9fa" rx="8"/>
  <text x="150" y="22" text-anchor="middle" font-size="13" font-weight="bold" fill="{COLOR_GRAY}">3 — Ölheber / Abreisbogen (Gooseneck)</text>
  <!-- Rohr von unten einkommend -->
  <line x1="90" y1="340" x2="90" y2="240" stroke="{COLOR_BLUE}" stroke-width="5"/>
  <!-- Aufwärtsbogen links -->
  <path d="M 90 240 Q 90 200 120 200 Q 150 200 150 240 L 150 270" fill="none" stroke="{COLOR_BLUE}" stroke-width="5"/>
  <!-- Linker Arm aufwärts -->
  <line x1="150" y1="270" x2="150" y2="120" stroke="{COLOR_BLUE}" stroke-width="5"/>
  <!-- Halbkreis oben (Scheitel) -->
  <path d="M 150 120 Q 150 90 180 90 Q 210 90 210 120 L 210 160" fill="none" stroke="{COLOR_BLUE}" stroke-width="5"/>
  <!-- Rechter Arm nach unten, dann horizontal -->
  <line x1="210" y1="160" x2="210" y2="200" stroke="{COLOR_BLUE}" stroke-width="5"/>
  <line x1="210" y1="200" x2="280" y2="200" stroke="{COLOR_BLUE}" stroke-width="5"/>
  <text x="270" y="195" font-size="11" fill="{COLOR_BLUE}">→</text>
  <!-- Öl-Sohlen markieren -->
  <ellipse cx="120" cy="244" rx="12" ry="5" fill="#8B4513" opacity="0.8"/>
  <text x="120" y="260" text-anchor="middle" font-size="8" fill="#8B4513">Öl-Sohle</text>
  <ellipse cx="180" cy="94" rx="12" ry="5" fill="#8B4513" opacity="0.8"/>
  <text x="180" y="82" text-anchor="middle" font-size="8" fill="#8B4513">Öl-Sohle</text>
  <!-- Pfeile -->
  <text x="82" y="338" font-size="12" fill="{COLOR_BLUE}">↑</text>
  <!-- Erklärung -->
  <text x="20" y="310" font-size="9" fill="{COLOR_GRAY}">Rohr kommt von unten → S-Bogen → langer Aufstieg</text>
  <text x="20" y="323" font-size="9" fill="{COLOR_GRAY}">→ enger U-Bogen oben → waagrecht ab (min. 1% Gefälle)</text>
  <text x="20" y="336" font-size="9" fill="{COLOR_GRAY}">Öl-Sohlen verhindern Rückfluss in die Steigleitung</text>
</svg>"""

    # Schema 4: Zusammenführung oben (beide Abreisbögen)
    svg4 = f"""<svg {svg_css} viewBox="0 0 320 380" xmlns="http://www.w3.org/2000/svg">
  <rect width="320" height="380" fill="#f8f9fa" rx="8"/>
  <text x="160" y="22" text-anchor="middle" font-size="13" font-weight="bold" fill="{COLOR_GRAY}">4 — Zusammenführung oben (beide Abreisbögen)</text>
  <!-- Speed Riser kommt von unten links -->
  <line x1="70" y1="340" x2="70" y2="220" stroke="{COLOR_BLUE}" stroke-width="4"/>
  <!-- Speed Riser Gooseneck (a) -->
  <path d="M 70 220 Q 70 190 90 190 Q 110 190 110 220 L 110 250" fill="none" stroke="{COLOR_BLUE}" stroke-width="4"/>
  <line x1="110" y1="250" x2="110" y2="160" stroke="{COLOR_BLUE}" stroke-width="4"/>
  <path d="M 110 160 Q 110 135 130 135 Q 150 135 150 160 L 150 185" fill="none" stroke="{COLOR_BLUE}" stroke-width="4"/>
  <line x1="150" y1="185" x2="160" y2="185" stroke="{COLOR_BLUE}" stroke-width="4"/>
  <!-- Öl-Sohlen Speed Riser -->
  <ellipse cx="90" cy="224" rx="10" ry="4" fill="#8B4513" opacity="0.8"/>
  <ellipse cx="130" cy="139" rx="10" ry="4" fill="#8B4513" opacity="0.8"/>
  <text x="70" y="130" font-size="8" fill="{COLOR_BLUE}">(a) Speed Riser</text>

  <!-- Main Riser kommt von unten rechts -->
  <line x1="250" y1="340" x2="250" y2="220" stroke="{COLOR_RED}" stroke-width="7"/>
  <!-- Main Riser Gooseneck (b) -->
  <path d="M 250 220 Q 250 190 230 190 Q 210 190 210 220 L 210 250" fill="none" stroke="{COLOR_RED}" stroke-width="7"/>
  <line x1="210" y1="250" x2="210" y2="160" stroke="{COLOR_RED}" stroke-width="7"/>
  <path d="M 210 160 Q 210 135 190 135 Q 170 135 170 160 L 170 185" fill="none" stroke="{COLOR_RED}" stroke-width="7"/>
  <line x1="170" y1="185" x2="160" y2="185" stroke="{COLOR_RED}" stroke-width="7"/>
  <!-- Öl-Sohlen Main Riser -->
  <ellipse cx="230" cy="224" rx="12" ry="5" fill="#8B4513" opacity="0.8"/>
  <ellipse cx="190" cy="139" rx="12" ry="5" fill="#8B4513" opacity="0.8"/>
  <text x="210" y="130" font-size="8" fill="{COLOR_RED}">(b) Main Riser</text>

  <!-- T-Stück in der Mitte -->
  <rect x="152" y="178" width="16" height="14" fill="{COLOR_GRAY}" rx="3"/>
  <line x1="160" y1="178" x2="160" y2="60" stroke="{COLOR_GRAY}" stroke-width="6"/>
  <text x="160" y="50" text-anchor="middle" font-size="10" fill="{COLOR_GRAY}">↑ zum Verdichter</text>

  <!-- Legende -->
  <text x="10" y="355" font-size="9" fill="#555">Öl-Sohlen (▓) verhindern: a) Rückfluss Speed Riser  b) Rückfluss Main Riser</text>
  <text x="10" y="368" font-size="9" fill="#555">Mindest-Gefälle Abgang: 1 % zum T-Stück hin</text>
</svg>"""

    tab1, tab2, tab3, tab4 = st.tabs(["1 Gesamt", "2 Ölfalle", "3 Abreisbogen", "4 Zusammenführung"])
    with tab1:
        st.markdown(svg1, unsafe_allow_html=True)
        st.caption("Gesamtbild: Speed Riser (dünn, links) + Main Riser (dick, rechts) mit Ölfallen und Abreisbögen (a+b) an der Zusammenführung.")
    with tab2:
        st.markdown(svg2, unsafe_allow_html=True)
        st.caption("U-Bogen (Ölfalle) am Fuß des Main Risers: Öl sammelt sich bei Teillast → sperrt Main Riser hydraulisch → Speed Riser übernimmt.")
    with tab3:
        st.markdown(svg3, unsafe_allow_html=True)
        st.caption("Gooseneck/Abreisbogen: S-förmiger Aufstieg → enger U-Bogen oben → horizontaler Abgang mit 1% Gefälle. Je 2× am Speed Riser im Abstand 3–4 m.")
    with tab4:
        st.markdown(svg4, unsafe_allow_html=True)
        st.caption("Zusammenführung oben: Beide Rohre mit eigenem Abreisbogen → T-Stück. (a) = Speed Riser Abreisbogen, (b) = Main Riser Abreisbogen.")

# ─────────────────────────────────────────────────────────────────────────────
# HAUPTBERECHNUNG & ERGEBNISANZEIGE
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.calculated:

    ref = REFRIGERANTS[ref_key]

    # --- Sättigungseigenschaften ---
    props0 = get_sat_props(ref_key, t0)  # Verdampfungsbedingungen
    propsc = get_sat_props(ref_key, tc)  # Verflüssigungsbedingungen

    p0    = props0["p_bar"]
    pc    = propsc["p_bar"]
    rho_v0 = props0["rho_v"]
    rho_vc = propsc["rho_v"]
    rho_l  = propsc["rho_l"]
    h_fg0  = props0["h_fg"]
    mu_v0  = props0["mu_v"]
    mu_vc  = propsc["mu_v"]

    # --- Massenstrom ---
    m_dot = Q_kW / h_fg0  # kg/s

    # --- Äquivalente Längen ---
    # Repräsentatives OD für Formstücke: 42mm (Mittelwert; wird unten pipe-spezifisch verfeinert)
    L_SL_total = L_SL_h + abs(L_SL_v)
    L_SL_eq = equiv_length(L_SL_total, n_el, n_bv, n_sv, 42)
    L_DL_eq  = equiv_length(L_DL, n_el, n_bv, 0,    42)
    L_FL_eq  = equiv_length(L_FL, n_el, n_bv, n_sv, 28)
    L_KL_eq  = equiv_length(L_KL, 2, 1, 0, 22)

    # --- v-Grenzen ---
    if app_code == "NK":
        vmin_sl_h  = 3.8;  vmax_sl = 18.0
        vmin_sl_v  = 7.6   # Steigleitung
        vmin_dl    = 5.0;  vmax_dl = 15.0
        vmin_dl_v  = 8.0
    else:
        vmin_sl_h  = 4.5;  vmax_sl = 18.0
        vmin_sl_v  = 9.0
        vmin_dl    = 5.0;  vmax_dl = 15.0
        vmin_dl_v  = 10.0

    vmin_fl = 0.5; vmax_fl = 1.5
    vmin_kl = 0.3; vmax_kl = 1.0

    # --- Druckverlustgrenzen ---
    dp_max_SL = dp_limit_K("SL", L_SL_total, app_code)
    dp_max_DL = dp_limit_K("DL", L_DL, app_code)
    dp_max_FL = dp_limit_K("FL", L_FL, app_code)
    dp_max_KL = dp_limit_K("FL", L_KL, app_code)

    # --- Hydrostatik ---
    dp_hydro_SL = hydrostatic_dp(rho_v0, h_SL)     # Pa – Steigleitung (Gas)
    dp_hydro_FL = hydrostatic_dp(rho_l,  h_FL)      # Pa – Flüssigkeit (neg = Gewinn)

    # --- Rohrauswahl Auto ---
    # Saugleitung: worst-case v_min = Steigleitung
    v_min_sl_combined = vmin_sl_v if L_SL_v > 0.5 else vmin_sl_h
    auto_SL, warns_SL = select_pipe(m_dot, rho_v0, mu_v0, L_SL_eq,
                                     v_min_sl_combined, vmax_sl, dp_max_SL,
                                     t0, h_fg0, rho_v0)
    auto_DL, warns_DL = select_pipe(m_dot, rho_vc, mu_vc, L_DL_eq,
                                     vmin_dl_v if L_DL > 1 else vmin_dl,
                                     vmax_dl, dp_max_DL,
                                     tc, propsc["h_fg"], rho_vc)
    auto_FL, warns_FL = select_pipe(m_dot, rho_l,  propsc["mu_v"]*1e-6, L_FL_eq,
                                     vmin_fl, vmax_fl, dp_max_FL,
                                     tc, propsc["h_fg"], rho_vc)
    auto_KL, warns_KL = select_pipe(m_dot, rho_l,  propsc["mu_v"]*1e-6, L_KL_eq,
                                     vmin_kl, vmax_kl, dp_max_KL,
                                     tc, propsc["h_fg"], rho_vc)

    # DSR prüfen: wenn v < v_min_v in Saugleitung
    sl_pipe_cur = CU_PIPES[max(0, min(len(CU_PIPES)-1, auto_SL + st.session_state.offsets["SL"]))]
    r_sl_test = calc_pipe(sl_pipe_cur, m_dot, rho_v0, mu_v0, L_SL_eq, t0, h_fg0, rho_v0)
    need_dsr_sl = (r_sl_test["v"] < vmin_sl_v and L_SL_v > 0.5)

    dl_pipe_cur = CU_PIPES[max(0, min(len(CU_PIPES)-1, auto_DL + st.session_state.offsets["DL"]))]
    r_dl_test = calc_pipe(dl_pipe_cur, m_dot, rho_vc, mu_vc, L_DL_eq, tc, propsc["h_fg"], rho_vc)
    need_dsr_dl = (r_dl_test["v"] < vmin_dl_v and L_DL > 1)

    # ── KOPF-METRIKEN ──────────────────────────────────────────────────────
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown(f"<div class='card-title'>⚙️ Anlagendaten — {ref['name']} | {app_code} | Q₀ = {Q_kW:.1f} kW</div>", unsafe_allow_html=True)

    # Warnung Kältemittel
    if ref["co2"]:
        st.markdown(f"<div class='err-box'>🔴 {ref['warning']}</div>", unsafe_allow_html=True)
    elif ref["a2l"]:
        st.markdown(f"<div class='warn-box'>⚠️ {ref['warning']}</div>", unsafe_allow_html=True)

    c1, c2, c3, c4, c5 = st.columns(5)
    def metric(col, label, val):
        col.markdown(f"<div class='metric-box'><div class='metric-val'>{val}</div><div class='metric-lbl'>{label}</div></div>", unsafe_allow_html=True)

    metric(c1, "Massenstrom ṁ", f"{m_dot*3600:.1f} kg/h")
    metric(c2, "p₀ Verdampfer", f"{p0:.2f} bar")
    metric(c3, "pc Kondensator", f"{pc:.2f} bar")
    metric(c4, "ρ_v Saugleitung", f"{rho_v0:.2f} kg/m³")
    metric(c5, "h_fg bei t₀", f"{h_fg0:.0f} kJ/kg")
    st.markdown("</div>", unsafe_allow_html=True)

    # ── LEITUNGEN ──────────────────────────────────────────────────────────
    tab_sl, tab_dl, tab_fl, tab_kl = st.tabs([
        "🔵 Saugleitung (SL)", "🔴 Druckleitung (DL)",
        "🟢 Flüssigkeitsleitung (FL)", "🟡 Kondensatleitung (KL)"
    ])

    with tab_sl:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        r_sl, dp_K_sl = show_pipe_result(
            f"Saugleitung — L_h={L_SL_h:.0f} m + L_v={L_SL_v:.0f} m | L_eq≈{L_SL_eq:.1f} m",
            "SL", auto_SL, m_dot, rho_v0, mu_v0, L_SL_eq,
            v_min_sl_combined, vmax_sl, dp_max_SL, t0, h_fg0, rho_v0,
            dp_hydro_Pa=dp_hydro_SL, warns=warns_SL
        )
        if need_dsr_sl:
            st.markdown(f"<div class='info-box'>ℹ️ Geschwindigkeit {r_sl_test['v']:.1f} m/s < v_min Steigleitung {vmin_sl_v:.1f} m/s → <b>Doppelsteigrohr (DSR) empfohlen!</b><br>Aktiviere DSR unten.</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with tab_dl:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        r_dl, dp_K_dl = show_pipe_result(
            f"Druckleitung — L={L_DL:.0f} m | L_eq≈{L_DL_eq:.1f} m",
            "DL", auto_DL, m_dot, rho_vc, mu_vc, L_DL_eq,
            vmin_dl_v if L_DL > 1 else vmin_dl, vmax_dl,
            dp_max_DL, tc, propsc["h_fg"], rho_vc,
            warns=warns_DL
        )
        if need_dsr_dl:
            st.markdown(f"<div class='info-box'>ℹ️ Geschwindigkeit zu niedrig → <b>Doppelsteigrohr für DL empfohlen!</b></div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with tab_fl:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        show_pipe_result(
            f"Flüssigkeitsleitung — L={L_FL:.0f} m | L_eq≈{L_FL_eq:.1f} m",
            "FL", auto_FL, m_dot, rho_l, propsc["mu_v"]*1e-6, L_FL_eq,
            vmin_fl, vmax_fl, dp_max_FL, tc, propsc["h_fg"], rho_vc,
            dp_hydro_Pa=dp_hydro_FL, warns=warns_FL
        )
        if h_FL < 0:
            gain_bar = abs(dp_hydro_FL) / 1e5
            st.markdown(f"<div class='info-box'>✅ Flüssigkeitsleitung fällt um {abs(h_FL):.1f} m → hydrostatischer Druckgewinn: {gain_bar*1000:.1f} mbar</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with tab_kl:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        show_pipe_result(
            f"Kondensatleitung — L={L_KL:.0f} m",
            "KL", auto_KL, m_dot, rho_l, propsc["mu_v"]*1e-6, L_KL_eq,
            vmin_kl, vmax_kl, dp_max_KL, tc, propsc["h_fg"], rho_vc,
            warns=warns_KL
        )
        st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# DSR – DOPPELSTEIGROHR (Double Suction / Discharge Riser)
# Erscheint NUR nach erfolgreicher Berechnung wenn v < v_min_vertikal
# ─────────────────────────────────────────────────────────────────────────────
    if need_dsr_sl or need_dsr_dl:
        st.markdown("---")
        st.markdown(f"<div style='font-size:22px;font-weight:bold;color:white;margin:10px 0 6px;'>🔀 Doppelsteigrohr-Auslegung (DSR)</div>", unsafe_allow_html=True)
        st.markdown(f"<div style='color:rgba(255,255,255,0.85);font-size:13px;margin-bottom:12px;'>Das Einzelrohr erreicht die Mindestgeschwindigkeit in der Steigleitung nicht — ein Doppelsteigrohr (Double Suction Riser) sichert Ölrückführung bei Teil- und Volllast.</div>", unsafe_allow_html=True)

    if need_dsr_sl:
        with st.expander("🔵 DSR Saugleitung aktivieren", expanded=st.session_state.dsr_sl):
            col_tog, _ = st.columns([2, 5])
            if col_tog.button("▶ DSR aktivieren" if not st.session_state.dsr_sl else "⏹ DSR deaktivieren",
                              key="dsr_sl_toggle"):
                st.session_state.dsr_sl = not st.session_state.dsr_sl
                st.rerun()

            if st.session_state.dsr_sl:
                st.markdown("<div class='card'>", unsafe_allow_html=True)
                st.markdown("<div class='card-title'>DSR Saugleitung — Dimensionierung</div>", unsafe_allow_html=True)

                # Speed Riser: kleinstes Rohr bei dem v(m_total) >= v_min_v
                speed_idx_sl = len(CU_PIPES) - 1
                for i, p in enumerate(CU_PIPES):
                    d_m = p["id"] / 1000.0
                    A   = math.pi * (d_m/2)**2
                    v_s = m_dot / (rho_v0 * A)
                    if v_s >= vmin_sl_v:
                        speed_idx_sl = i
                        break

                # Main Riser: 2 Größen über Speed Riser (Vollast, beide offen)
                main_idx_sl  = min(len(CU_PIPES) - 1, speed_idx_sl + 2)

                # Manuelle Overrides mit Offset
                off_speed = st.session_state.offsets.get("SL_speed", 0)
                off_main  = st.session_state.offsets.get("SL_main",  0)
                cur_speed = max(0, min(len(CU_PIPES)-1, speed_idx_sl + off_speed))
                cur_main  = max(0, min(len(CU_PIPES)-1, main_idx_sl  + off_main))

                c_sp, c_ma = st.columns(2)
                with c_sp:
                    st.markdown("**Speed Riser (kleines Rohr)**")
                    st.caption("Teillast: ganzer Massenstrom → v ≥ v_min")
                    b1, lbl1, b2 = st.columns([1,4,1])
                    if b1.button("−", key="dsr_sl_speed_m"):
                        st.session_state.offsets["SL_speed"] = max(-speed_idx_sl, off_speed-1); st.rerun()
                    if b2.button("+", key="dsr_sl_speed_p"):
                        st.session_state.offsets["SL_speed"] = min(len(CU_PIPES)-1-speed_idx_sl, off_speed+1); st.rerun()
                    p_sp = CU_PIPES[cur_speed]
                    lbl1.markdown(f"<div style='text-align:center;padding-top:6px;font-size:18px;font-weight:bold;color:{COLOR_GRAY};'>⌀ {p_sp['od']:.0f} / {p_sp['id']:.0f} mm</div>", unsafe_allow_html=True)
                    r_sp = calc_pipe(p_sp, m_dot, rho_v0, mu_v0, L_SL_v, t0, h_fg0, rho_v0)
                    st.markdown(velocity_bar_html(r_sp["v"], vmin_sl_v, vmax_sl, "Speed Riser Teillast"), unsafe_allow_html=True)
                    st.markdown(f"**ṁ** = {m_dot*3600:.1f} kg/h | **Δp** = {r_sp['dp_K']:.3f} K | {r_sp['dp_bar']*1000:.1f} mbar")

                with c_ma:
                    st.markdown("**Main Riser (großes Rohr)**")
                    st.caption("Volllast: Öl-Siphon offen, beide Rohre aktiv")
                    b3, lbl2, b4 = st.columns([1,4,1])
                    if b3.button("−", key="dsr_sl_main_m"):
                        st.session_state.offsets["SL_main"] = max(-main_idx_sl, off_main-1); st.rerun()
                    if b4.button("+", key="dsr_sl_main_p"):
                        st.session_state.offsets["SL_main"] = min(len(CU_PIPES)-1-main_idx_sl, off_main+1); st.rerun()
                    p_ma = CU_PIPES[cur_main]
                    lbl2.markdown(f"<div style='text-align:center;padding-top:6px;font-size:18px;font-weight:bold;color:{COLOR_GRAY};'>⌀ {p_ma['od']:.0f} / {p_ma['id']:.0f} mm</div>", unsafe_allow_html=True)
                    r_ma = calc_pipe(p_ma, m_dot, rho_v0, mu_v0, L_SL_v, t0, h_fg0, rho_v0)
                    # Parallelberechnung (vereinfacht: v proportional Querschnitt)
                    A_sp = math.pi * (p_sp["id"]/2000)**2
                    A_ma = math.pi * (p_ma["id"]/2000)**2
                    m_main_full = m_dot * A_ma / (A_sp + A_ma)
                    m_speed_full= m_dot * A_sp / (A_sp + A_ma)
                    r_ma_full = calc_pipe(p_ma, m_main_full, rho_v0, mu_v0, L_SL_v, t0, h_fg0, rho_v0)
                    st.markdown(velocity_bar_html(r_ma_full["v"], 0, vmax_sl, "Main Riser Volllast"), unsafe_allow_html=True)
                    st.markdown(f"**ṁ_main** = {m_main_full*3600:.1f} kg/h | **Δp** = {r_ma_full['dp_K']:.3f} K | {r_ma_full['dp_bar']*1000:.1f} mbar")

                st.info(f"**Funktionsprinzip:** Bei Teillast → Öl füllt Ölheber im Main Riser → sperrt diesen ab → aller Massenstrom durch Speed Riser ≥ v_min = {vmin_sl_v} m/s. Bei Volllast → Druck überwindet Öl-Siphon → beide Rohre offen → Gesamtgeschwindigkeit = {(m_dot/(rho_v0*(A_sp+A_ma))):.1f} m/s.", icon="ℹ️")
                st.markdown("</div>", unsafe_allow_html=True)

                # Schema-Zeichnung
                if st.button("📐 Schaltungsschema DSR anzeigen", key="schema_sl"):
                    show_dsr_schema()

    if need_dsr_dl:
        with st.expander("🔴 DSR Druckleitung aktivieren", expanded=st.session_state.dsr_dl):
            col_tog2, _ = st.columns([2, 5])
            if col_tog2.button("▶ DSR DL aktivieren" if not st.session_state.dsr_dl else "⏹ DSR DL deaktivieren",
                               key="dsr_dl_toggle"):
                st.session_state.dsr_dl = not st.session_state.dsr_dl
                st.rerun()

            if st.session_state.dsr_dl:
                st.markdown("<div class='card'>", unsafe_allow_html=True)
                st.markdown("<div class='card-title'>DSR Druckleitung — Dimensionierung</div>", unsafe_allow_html=True)

                speed_idx_dl = len(CU_PIPES) - 1
                for i, p in enumerate(CU_PIPES):
                    d_m = p["id"] / 1000.0
                    A   = math.pi * (d_m/2)**2
                    v_s = m_dot / (rho_vc * A)
                    if v_s >= vmin_dl_v:
                        speed_idx_dl = i
                        break
                main_idx_dl = min(len(CU_PIPES) - 1, speed_idx_dl + 2)

                off_speed_dl = st.session_state.offsets.get("DL_speed", 0)
                off_main_dl  = st.session_state.offsets.get("DL_main",  0)
                cur_speed_dl = max(0, min(len(CU_PIPES)-1, speed_idx_dl + off_speed_dl))
                cur_main_dl  = max(0, min(len(CU_PIPES)-1, main_idx_dl  + off_main_dl))

                c_sp2, c_ma2 = st.columns(2)
                with c_sp2:
                    st.markdown("**Speed Riser DL**")
                    b5, lbl3, b6 = st.columns([1,4,1])
                    if b5.button("−", key="dsr_dl_speed_m"):
                        st.session_state.offsets["DL_speed"] = max(-speed_idx_dl, off_speed_dl-1); st.rerun()
                    if b6.button("+", key="dsr_dl_speed_p"):
                        st.session_state.offsets["DL_speed"] = min(len(CU_PIPES)-1-speed_idx_dl, off_speed_dl+1); st.rerun()
                    p_sp2 = CU_PIPES[cur_speed_dl]
                    lbl3.markdown(f"<div style='text-align:center;padding-top:6px;font-size:18px;font-weight:bold;color:{COLOR_GRAY};'>⌀ {p_sp2['od']:.0f} / {p_sp2['id']:.0f} mm</div>", unsafe_allow_html=True)
                    r_sp2 = calc_pipe(p_sp2, m_dot, rho_vc, mu_vc, L_DL, tc, propsc["h_fg"], rho_vc)
                    st.markdown(velocity_bar_html(r_sp2["v"], vmin_dl_v, vmax_dl, "Speed Riser Teillast"), unsafe_allow_html=True)

                with c_ma2:
                    st.markdown("**Main Riser DL**")
                    b7, lbl4, b8 = st.columns([1,4,1])
                    if b7.button("−", key="dsr_dl_main_m"):
                        st.session_state.offsets["DL_main"] = max(-main_idx_dl, off_main_dl-1); st.rerun()
                    if b8.button("+", key="dsr_dl_main_p"):
                        st.session_state.offsets["DL_main"] = min(len(CU_PIPES)-1-main_idx_dl, off_main_dl+1); st.rerun()
                    p_ma2 = CU_PIPES[cur_main_dl]
                    lbl4.markdown(f"<div style='text-align:center;padding-top:6px;font-size:18px;font-weight:bold;color:{COLOR_GRAY};'>⌀ {p_ma2['od']:.0f} / {p_ma2['id']:.0f} mm</div>", unsafe_allow_html=True)
                    A_sp2 = math.pi*(p_sp2["id"]/2000)**2; A_ma2 = math.pi*(p_ma2["id"]/2000)**2
                    m_ma2 = m_dot * A_ma2/(A_sp2+A_ma2)
                    r_ma2_full = calc_pipe(p_ma2, m_ma2, rho_vc, mu_vc, L_DL, tc, propsc["h_fg"], rho_vc)
                    st.markdown(velocity_bar_html(r_ma2_full["v"], 0, vmax_dl, "Main Riser Volllast"), unsafe_allow_html=True)

                st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL-EXPORT
# ─────────────────────────────────────────────────────────────────────────────
    # Excel Export Button
    st.markdown("---")
    if st.button("📊 Excel-Nachweis generieren", use_container_width=True):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Projektdaten
            proj_data = {
                "Feld": ["Kunde", "Projekt", "Bearbeiter", "Datum",
                         "Kältemittel", "Anwendung", "t₀ (°C)", "tc (°C)",
                         "Q₀ (kW)", "Massenstrom (kg/h)"],
                "Wert": [kunde, projekt, bearbeiter, datetime.now().strftime("%d.%m.%Y"),
                         ref["name"], app_code, t0, tc,
                         Q_kW, round(m_dot*3600, 2)]
            }
            df_proj = pd.DataFrame(proj_data)
            df_proj.to_excel(writer, sheet_name="Projekt", index=False)

            # Ergebnisse pro Leitung
            def pipe_row(label, line_code, auto_i, m, rho, mu, L_eq, vmin, vmax, dp_max, T_sat, h_fg_val, rho_v_val, dp_hydro=0):
                off = st.session_state.offsets.get(line_code, 0)
                cur = max(0, min(len(CU_PIPES)-1, auto_i + off))
                p = CU_PIPES[cur]
                r = calc_pipe(p, m, rho, mu, L_eq, T_sat, h_fg_val, rho_v_val)
                dp_total = (r["dp_Pa"] + dp_hydro) / 1e5
                T_K = T_sat + 273.15
                dp_dT = (h_fg_val * 1000 * rho_v_val) / T_K
                dp_K_val = (r["dp_Pa"] + dp_hydro) / dp_dT if dp_dT > 0 else r["dp_K"]
                insul = insulation_thickness_mm(T_sat, T_amb, phi, p["od"])
                status = "OK" if (vmin <= r["v"] <= vmax and dp_K_val <= dp_max) else "PRÜFEN"
                return {
                    "Leitung": label,
                    "OD (mm)": p["od"], "ID (mm)": p["id"], "Wandstärke (mm)": p["wall"],
                    "L_eq (m)": round(L_eq, 2),
                    "ṁ (kg/h)": round(m*3600, 2),
                    "ρ (kg/m³)": round(rho, 3),
                    "v (m/s)": round(r["v"], 3),
                    "v_min (m/s)": vmin, "v_max (m/s)": vmax,
                    "Δp_Reib (K)": round(r["dp_K"], 4),
                    "Δp_gesamt (K)": round(dp_K_val, 4),
                    "Δp_gesamt (bar)": round(dp_total, 5),
                    "Grenzwert (K)": dp_max,
                    "Isolierung (mm)": insul,
                    "Status": status
                }

            rows = [
                pipe_row("Saugleitung", "SL", auto_SL, m_dot, rho_v0, mu_v0, L_SL_eq,
                         v_min_sl_combined, vmax_sl, dp_max_SL, t0, h_fg0, rho_v0, dp_hydro_SL),
                pipe_row("Druckleitung", "DL", auto_DL, m_dot, rho_vc, mu_vc, L_DL_eq,
                         vmin_dl, vmax_dl, dp_max_DL, tc, propsc["h_fg"], rho_vc),
                pipe_row("Flüssigkeitsleitung", "FL", auto_FL, m_dot, rho_l, propsc["mu_v"]*1e-6, L_FL_eq,
                         vmin_fl, vmax_fl, dp_max_FL, tc, propsc["h_fg"], rho_vc, dp_hydro_FL),
                pipe_row("Kondensatleitung", "KL", auto_KL, m_dot, rho_l, propsc["mu_v"]*1e-6, L_KL_eq,
                         vmin_kl, vmax_kl, dp_max_KL, tc, propsc["h_fg"], rho_vc),
            ]
            df_res = pd.DataFrame(rows)
            df_res.to_excel(writer, sheet_name="Rohrdimensionierung", index=False)

            # Formatierung
            wb = writer.book
            for sh_name in ["Projekt", "Rohrdimensionierung"]:
                ws = wb[sh_name]
                blue_fill = PatternFill(start_color="36A9E1", end_color="36A9E1", fill_type="solid")
                gray_fill = PatternFill(start_color="3C3C3B", end_color="3C3C3B", fill_type="solid")
                white_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
                for cell in ws[1]:
                    cell.fill = gray_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal="center")
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")
                        if sh_name == "Rohrdimensionierung" and cell.column == df_res.columns.get_loc("Status") + 1:
                            if cell.value == "OK":
                                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                                cell.font = Font(color="FFFFFF", bold=True, name="Arial")
                            elif cell.value == "PRÜFEN":
                                cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                                cell.font = Font(color="FFFFFF", bold=True, name="Arial")
                for col in ws.columns:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 3, 30))

        buf.seek(0)
        st.download_button(
            label="💾 Excel herunterladen",
            data=buf,
            file_name=f"coolROHR_{projekt.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:
    # Startbildschirm wenn noch nicht berechnet
    st.markdown(f"""
<div style='background:white;border-radius:12px;padding:40px;text-align:center;margin-top:30px;'>
  <div style='font-size:48px;margin-bottom:16px;'>🔧</div>
  <div style='font-size:22px;font-weight:bold;color:{COLOR_GRAY};margin-bottom:10px;'>°coolROHR — Kältemittel-Rohrdimensionierung</div>
  <div style='font-size:14px;color:#666;max-width:500px;margin:0 auto;line-height:1.7;'>
    Eingaben links in der Sidebar ausfüllen, dann <b>⚡ BERECHNEN</b> drücken.<br><br>
    Unterstützte Kältemittel: R32, R744 (CO₂), R1234yf, R455A, R452A, R513A, R449A<br>
    Leitungstypen: Saug-, Druck-, Flüssigkeits- und Kondensatleitung<br>
    DSR (Doppelsteigrohr) wird automatisch vorgeschlagen wenn v &lt; v_min
  </div>
  <div style='margin-top:24px;font-size:12px;color:#aaa;'>°coolsulting e.U. | Michael Schäpers | v2.0</div>
</div>""", unsafe_allow_html=True)

